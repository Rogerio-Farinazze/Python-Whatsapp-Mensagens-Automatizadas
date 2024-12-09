import locale
import sys
import os
import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QPushButton,
    QTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
    QLabel,
    QDesktopWidget,
    QCheckBox,
)
# Define o locale para o padrão brasileiro
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

class JanelaPrincipal(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

        # Atributos para armazenar dados
        self.dados_planilha = []
        self.caminho_planilha = ""
        self.carregando_tabela = False  # Evita loops desnecessários ao alterar o grid


    def init_ui(self):
        sizeDisplay = QDesktopWidget().screenGeometry(-1)
        sizeDisplayX = sizeDisplay.width()
        sizeDisplayY = sizeDisplay.height()
        sizeGeometryX = 800
        sizeGeometryY = 600
        positionX = int(sizeDisplayX/2 - sizeGeometryX/2)
        positionY = int(sizeDisplayY/2 - sizeGeometryY/2)

        self.setWindowTitle("Envio de Mensagens WhatsApp")
        self.setGeometry(positionX, positionY, sizeGeometryX, sizeGeometryY)

        # Layout principal
        layout = QVBoxLayout()

        # Botão "Baixar Planilha"
        botao_download = QPushButton("Baixar Planilha Padrão")
        botao_download.clicked.connect(self.baixar_planilha)
        layout.addWidget(botao_download)

        # Botão para upload da planilha
        self.botao_upload = QPushButton("Fazer Upload da Planilha")
        self.botao_upload.clicked.connect(self.carregar_planilha)
        layout.addWidget(self.botao_upload)

        # Tabela para exibir os dados da planilha
        self.tabela = QTableWidget()
        self.tabela.setColumnCount(5)  # Número de colunas: Nome, Telefone, Data Limite, Valor, Enviar?
        self.tabela.setHorizontalHeaderLabels(["Nome", "Telefone", "Data Limite", "Valor", "Enviar?"])
        self.tabela.cellChanged.connect(self.atualizar_dados)  # Detecta alterações no grid
        layout.addWidget(self.tabela)

        # Botão para remover linha
        self.botao_remover = QPushButton("Remover Linha")
        self.botao_remover.clicked.connect(self.remover_linha)
        layout.addWidget(self.botao_remover)

        # Campo para o usuário inserir a mensagem
        layout.addWidget(QLabel("Mensagem personalizada:"))
        self.campo_mensagem = QTextEdit("Olá {nome}, sua mensalidade no valor de R$ {valor_devido} vence no dia {data_limite}.")
        layout.addWidget(self.campo_mensagem)

        # Botão para iniciar o envio
        self.botao_enviar = QPushButton("Iniciar Envio")
        self.botao_enviar.clicked.connect(self.iniciar_envio)
        layout.addWidget(self.botao_enviar)

        # Configura o layout na janela principal
        self.setLayout(layout)

    def baixar_planilha(self):
        # Define o local para salvar o arquivo usando uma janela de seleção
        caminho, _ = QFileDialog.getSaveFileName(
            self, "Salvar Planilha", "planilha_padrao.xlsx", "Arquivos Excel (*.xlsx)"
        )

        if caminho:
            try:
                # Cria a planilha padrão
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Dados"
                ws.append(["Nome", "Telefone", "Data Limite", "Valor"])  # Cabeçalhos

                # Salva o arquivo no caminho especificado
                wb.save(caminho)

                # Exibe a mensagem de confirmação com a opção de abrir o arquivo
                self.mostrar_mensagem_download_concluido(caminho)
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao salvar a planilha: {e}")

    def mostrar_mensagem_download_concluido(self, caminho):
        # Cria a mensagem com botões personalizados
        mensagem = QMessageBox(self)
        mensagem.setIcon(QMessageBox.Information)
        mensagem.setWindowTitle("Download Concluído")
        mensagem.setText(f"Planilha padrão salva em:\n{caminho}")

        # Cria os botões manualmente para personalizar os textos
        botao_fechar = mensagem.addButton("Fechar", QMessageBox.RejectRole)
        botao_abrir = mensagem.addButton("Abrir Arquivo", QMessageBox.AcceptRole)

        # Mostra a mensagem
        mensagem.exec_()

        # Verifica qual botão foi clicado
        if mensagem.clickedButton() == botao_abrir:
            # Abre o arquivo no aplicativo padrão do sistema operacional
            os.startfile(caminho)  # Apenas para Windows; no Linux/Mac, usar subprocess.

    def carregar_planilha(self):
        # Seleciona o arquivo de planilha
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Arquivos Excel (*.xlsx)")
        if not caminho:
            return

        self.caminho_planilha = caminho
        try:
            # Carrega os dados da planilha
            workbook = openpyxl.load_workbook(caminho)
            pagina_clientes = workbook.active

            # Limpa dados anteriores
            self.dados_planilha.clear()
            self.tabela.setRowCount(0)

            # Lê as linhas da planilha e armazena em `dados_planilha`
            for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
                self.dados_planilha.append(list(linha) + [True])  # Adiciona a coluna "Enviar?"

            # Preenche os dados na tabela
            self.carregando_tabela = True  # Evita detecção de alterações durante o carregamento
            self.tabela.setRowCount(len(self.dados_planilha))
            for i, linha in enumerate(self.dados_planilha):
                for j, valor in enumerate(linha):
                    if j == 3:  # Coluna Valor
                        valor = f"R$ {locale.format_string('%.2f', valor, grouping=True)}" if valor is not None else "R$ 0,00"
                    elif j == 4:  # Coluna Enviar? (checkbox)
                        checkbox = QTableWidgetItem()
                        checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                        checkbox.setCheckState(Qt.Checked if valor else Qt.Unchecked)
                        self.tabela.setItem(i, j, checkbox)
                        continue
                    self.tabela.setItem(i, j, QTableWidgetItem(str(valor)))
            self.carregando_tabela = False

            QMessageBox.information(self, "Sucesso", "Planilha carregada com sucesso!")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar a planilha: {e}")

    def atualizar_dados(self, row, column):
        # Atualiza os dados na lista quando a tabela é alterada
        if self.carregando_tabela:
            return  # Ignora alterações durante o carregamento inicial

        if column == 4:  # Coluna Enviar? (checkbox)
            checkbox = self.tabela.item(row, column)
            self.dados_planilha[row][column] = checkbox.checkState() == Qt.Checked
        else:
            novo_valor = self.tabela.item(row, column).text()

            # Formata a coluna Valor
            if column == 3:
                novo_valor = novo_valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
                try:
                    valor_formatado = float(novo_valor)
                    self.dados_planilha[row][column] = valor_formatado
                    valor_exibicao = f"R$ {locale.format_string('%.2f', valor_formatado, grouping=True)}"
                    self.tabela.blockSignals(True)
                    self.tabela.setItem(row, column, QTableWidgetItem(valor_exibicao))
                    self.tabela.blockSignals(False)
                except ValueError:
                    QMessageBox.warning(self, "Erro", f"Valor inválido na linha {row + 1}.")
                    return
            else:
                self.dados_planilha[row][column] = novo_valor

    def remover_linha(self):
        # Remove a linha selecionada no grid
        linha_selecionada = self.tabela.currentRow()
        if linha_selecionada == -1:
            QMessageBox.warning(self, "Aviso", "Selecione uma linha para remover.")
            return

        # Remove a linha da tabela e da lista de dados
        self.tabela.removeRow(linha_selecionada)
        self.dados_planilha.pop(linha_selecionada)

    def iniciar_envio(self):
        # Filtra apenas as linhas marcadas para envio
        linhas_para_enviar = [linha for linha in self.dados_planilha if linha[4]]  # Verifica o estado do checkbox

        if not linhas_para_enviar:
            QMessageBox.warning(self, "Aviso", "Nenhuma linha marcada para envio.")
            return

        # Obtém a mensagem personalizada
        mensagem_base = self.campo_mensagem.toPlainText()
        if not mensagem_base:
            QMessageBox.warning(self, "Aviso", "Insira a mensagem personalizada antes de iniciar o envio.")
            return

        # Adiciona o rodapé à mensagem
        rodape = "\n\nSiga nas redes sociais @Conectadomogi\nAcesse o site https://colegioconectado.com.br"
        mensagem_base += rodape

        # Inicia o processo de envio
        for linha in linhas_para_enviar:
            nome, telefone, data_limite, valor_devido, _ = linha
            mensagem = mensagem_base.format(
                nome=nome, telefone=telefone, data_limite=data_limite, valor_devido=f"R$ {locale.format_string('%.2f', valor_devido, grouping=True)}" if valor_devido is not None else "R$ 0,00"
            )

            try:
                # Envia a mensagem pelo WhatsApp
                link_whats = f"https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}"
                webbrowser.open(link_whats)
                sleep(6)
                pyautogui.press('esc')
                sleep(2)
                pyautogui.press('enter')
                sleep(2)
                pyautogui.hotkey('ctrl', 'w')
                sleep(2)
            except Exception as e:
                print(f"Erro ao enviar mensagem para {nome}: {e}")
                with open('erros.csv', 'a+', newline='', encoding='utf-8') as arquivo:
                    arquivo.write(f'\r\n{nome},{telefone}')

        QMessageBox.information(self, "Envio Concluído", "Mensagens enviadas com sucesso!")

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Instancia a janela principal
    janela = JanelaPrincipal()
    janela.show()

    # Executa o loop da aplicação
    sys.exit(app.exec_())
